"""
API: Reportes
Exportacion de combates guardados en PDF y Excel.
Soporta filtros por campeonato/tatami y exportacion dividida (ZIP):
- Sin filtros: un archivo por campeonato.
- Con campeonato: un archivo por tatami.
"""
import io
import re
import zipfile
from datetime import datetime, timedelta, timezone
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from ..extensions import db
from ..models.usuario import Usuario
from ..models.combate import Combate
from ..models.tatami import Tatami, SesionTatami
from ..models.campeonato import Campeonato
from ..models.llave import Llave
from ..timeutil import a_local, iso_utc
from ..filei18n import trad, idioma_request
from .llaves import podio_llave

reportes_bp = Blueprint("reportes", __name__)

# Traductor por defecto (español) para llamadas internas sin idioma explícito.
_T_ES = trad("es")

# Clave de traducción de cada ronda (el texto vive en filei18n).
RONDA_CLAVE = {"r1": "ronda_r1", "r2": "ronda_r2", "oro": "ronda_oro", "figuras": "figuras"}


def _require_admin():
    from .scoping import require_admin
    return require_admin()


def _build_query(args, user=None):
    """Query de (Combate, Tatami, Campeonato) con filtros opcionales.
    Con `user` admin no-superadmin, se limita a SUS campeonatos (workspace)."""
    q = db.session.query(Combate, Tatami, Campeonato).join(
        SesionTatami, Combate.sesion_tatami_id == SesionTatami.id
    ).join(
        Tatami, SesionTatami.tatami_id == Tatami.id
    ).outerjoin(
        Campeonato, Tatami.campeonato_id == Campeonato.id
    )

    if user is not None and user.rol == "admin" and not user.es_super:
        q = q.filter(Campeonato.created_by == user.id)

    if args.get("campeonato_id"):
        q = q.filter(Tatami.campeonato_id == int(args["campeonato_id"]))
    if args.get("tatami_id"):
        q = q.filter(SesionTatami.tatami_id == int(args["tatami_id"]))
    # Selección puntual de registros: ids=1,5,9 (descarga individual)
    if args.get("ids"):
        try:
            ids = [int(x) for x in str(args["ids"]).split(",") if x.strip()]
            if ids:
                q = q.filter(Combate.id.in_(ids))
        except ValueError:
            pass
    if args.get("categoria_id"):
        q = q.filter(Combate.categoria_id == int(args["categoria_id"]))
    # Filtro por tipo: los registros de figuras siempre guardan
    # ronda_final = "figuras"; el resto son combates.
    if args.get("tipo") == "figuras":
        q = q.filter(Combate.ronda_final == "figuras")
    elif args.get("tipo") == "combate":
        q = q.filter(Combate.ronda_final != "figuras")
    # Los filtros de fecha llegan como días LOCALES ("2026-07-16") y la BD
    # guarda timestamps UTC: se convierte el límite local a UTC. Además,
    # "hasta" una fecha sin hora significa ese día COMPLETO (antes cortaba en
    # la medianoche y excluía casi todo el día seleccionado).
    if args.get("desde"):
        try:
            desde = datetime.fromisoformat(args["desde"])
            if desde.tzinfo is None:
                desde = desde.astimezone()  # medianoche local del servidor
            desde_utc = desde.astimezone(timezone.utc).replace(tzinfo=None)
            q = q.filter(Combate.created_at >= desde_utc)
        except Exception:
            pass
    if args.get("hasta"):
        try:
            hasta = datetime.fromisoformat(args["hasta"])
            if hasta.tzinfo is None:
                hasta = hasta.astimezone()
            if hasta.hour == 0 and hasta.minute == 0 and hasta.second == 0:
                hasta = hasta + timedelta(days=1)  # incluir el día completo
            hasta_utc = hasta.astimezone(timezone.utc).replace(tzinfo=None)
            q = q.filter(Combate.created_at < hasta_utc)
        except Exception:
            pass

    q = q.order_by(Combate.created_at.desc())
    return q


def _rows_filtradas(args, user=None):
    """
    Filas (combate, tatami, camp) con todos los filtros aplicados.
    El filtro de categoría se aplica en Python porque el nombre de la
    categoría vive dentro del JSON jueces_detalle.
    """
    rows = _build_query(args, user=user).all()
    categoria = (args.get("categoria") or "").strip().lower()
    if categoria:
        rows = [
            r for r in rows
            if _nombre_categoria_registro(r[0]).strip().lower() == categoria
        ]
    return rows


def _resumen_categorias(rows):
    """Conteo por nombre de categoría con su forma de puntuación."""
    resumen = {}
    for c, _t, _camp in rows:
        nombre = _nombre_categoria_registro(c)
        tipo = _tipo_registro(c)
        item = resumen.setdefault(nombre, {
            "nombre": nombre,
            # Forma de puntuación: "combate" (Hong vs Chung) o
            # "individual" (sistema de figuras reutilizado por categorías)
            "puntuacion": "individual" if tipo == "figuras" else "combate",
            "cantidad": 0,
        })
        item["cantidad"] += 1
    return sorted(resumen.values(), key=lambda x: (-x["cantidad"], x["nombre"]))


def _slug(texto, fallback="reporte"):
    """Texto seguro para nombres de archivo."""
    limpio = re.sub(r"[^A-Za-z0-9]+", "-", str(texto or "")).strip("-").lower()
    return limpio[:40] or fallback


def _contexto_filtros(args, t=None):
    """Filtros activos con nombres legibles (para títulos y nombre de archivo)."""
    t = t or _T_ES
    ctx = {
        "campeonato": None,
        "tatami": None,
        "categoria": None,
        "tipo": None,
        "tipo_slug": None,
        "num_ids": 0,
    }
    if args.get("campeonato_id"):
        camp = Campeonato.query.get(int(args["campeonato_id"]))
        if camp:
            ctx["campeonato"] = camp.nombre
    if args.get("tatami_id"):
        tatami = Tatami.query.get(int(args["tatami_id"]))
        if tatami:
            ctx["tatami"] = t("tatami_n", n=tatami.numero)
            if not ctx["campeonato"] and tatami.campeonato_id:
                camp = Campeonato.query.get(tatami.campeonato_id)
                if camp:
                    ctx["campeonato"] = camp.nombre
    if args.get("categoria"):
        ctx["categoria"] = str(args["categoria"]).strip()
    if args.get("tipo") == "combate":
        ctx["tipo"] = t("punt_combate")
        ctx["tipo_slug"] = "combate"
    elif args.get("tipo") == "figuras":
        ctx["tipo"] = t("punt_individual")
        ctx["tipo_slug"] = "figuras"
    if args.get("ids"):
        ctx["num_ids"] = len([x for x in str(args["ids"]).split(",") if x.strip()])
    return ctx


def _subtitulo_filtros(ctx, t=None):
    t = t or _T_ES
    partes = []
    if ctx.get("campeonato"):
        partes.append(t("sub_campeonato", v=ctx["campeonato"]))
    if ctx.get("tatami"):
        partes.append(ctx["tatami"])
    if ctx.get("categoria"):
        partes.append(t("sub_categoria", v=ctx["categoria"]))
    elif ctx.get("tipo"):
        partes.append(ctx["tipo"])
    if ctx.get("num_ids"):
        partes.append(t("sub_seleccion", n=ctx["num_ids"]))
    return " — ".join(partes)


def _nombre_archivo(ctx, ext):
    """
    Nombre de descarga específico para identificar el documento de un vistazo:
    dinamyt_<campeonato>_<tatami>_<categoría|tipo>_<selección-N>_<fecha_hora>.ext
    La fecha con hora y segundos evita sobrescrituras y duplicados.
    """
    partes = ["dinamyt"]
    if ctx.get("campeonato"):
        partes.append(_slug(ctx["campeonato"], "campeonato"))
    if ctx.get("tatami"):
        partes.append(_slug(ctx["tatami"], "tatami"))
    if ctx.get("categoria"):
        partes.append(_slug(ctx["categoria"], "categoria"))
    elif ctx.get("tipo_slug"):
        partes.append("combates" if ctx["tipo_slug"] == "combate" else "individual")
    if ctx.get("num_ids"):
        partes.append(f"seleccion-{ctx['num_ids']}")
    if len(partes) == 1:
        partes.append("resultados")
    partes.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
    return "_".join(partes) + f".{ext}"


def _format_momento_evento(entrada):
    """
    Formatea el timestamp real de una entrada del historial EN HORA LOCAL
    del servidor (la del evento). Los momentos se guardan en UTC; mostrarlos
    sin convertir dejaba los reportes 5 horas corridos respecto al reloj
    que vieron los jueces.
    """
    momento = entrada.get("momento") or entrada.get("ts")
    if not momento:
        return "-"
    if isinstance(momento, (int, float)):
        try:
            # fromtimestamp sin tz = hora local del servidor
            return datetime.fromtimestamp(momento / 1000).strftime("%d/%m/%Y %H:%M:%S")
        except Exception:
            return str(momento)
    if isinstance(momento, str):
        try:
            dt = datetime.fromisoformat(momento)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone().strftime("%d/%m/%Y %H:%M:%S")
        except ValueError:
            return momento.replace("T", " ")[:19]
    return str(momento)


def _fecha_local_txt(dt, formato="%d/%m/%Y %H:%M"):
    """Fecha guardada en UTC → texto en hora local del servidor."""
    local = a_local(dt)
    return local.strftime(formato) if local else "-"


def _jueces_meta(combate):
    detalle = combate.jueces_detalle or {}
    return detalle.get("asignaciones") or detalle.get("jueces_meta") or {}


def _juez_meta_para_evento(combate, entrada):
    rol = entrada.get("juez_rol") or entrada.get("juez")
    meta = _jueces_meta(combate).get(rol, {})
    return {
        "nombre": entrada.get("juez_nombre") or meta.get("nombre") or rol or "-",
        "email": entrada.get("juez_email") or meta.get("email") or "-",
        "asignacion": entrada.get("juez_asignacion") or meta.get("asignacion") or rol or "-",
        "rol": rol or "-",
        "acceso": entrada.get("juez_acceso") or meta.get("origen") or "-",
    }


def _jueces_list(combate):
    jueces = []
    for rol, meta in sorted(_jueces_meta(combate).items()):
        jueces.append({
            "rol_tatami": meta.get("rol_tatami") or rol,
            "asignacion": meta.get("asignacion") or rol,
            "nombre": meta.get("nombre") or "-",
            "email": meta.get("email") or "-",
            "origen": meta.get("origen") or "-",
            "asignado_at": meta.get("asignado_at"),
            "asignado_por": meta.get("asignado_por"),
        })
    return jueces


def _jueces_resumen(combate):
    partes = []
    for juez in _jueces_list(combate):
        email = juez["email"] if juez["email"] != "-" else ""
        # "pin" aparece solo en combates guardados antes de eliminar el acceso por PIN
        origen = "Asignado" if juez["origen"] == "asignacion" else "Directo"
        partes.append(
            f"{juez['asignacion']}: {juez['nombre']}"
            + (f" <{email}>" if email else "")
            + f" ({origen})"
        )
    return "; ".join(partes) if partes else "-"


def _detalle_registro(combate):
    return combate.jueces_detalle or {}


def _tipo_registro(combate):
    detalle = _detalle_registro(combate)
    if detalle.get("tipo"):
        return detalle["tipo"]
    if combate.categoria and combate.categoria.slug:
        return combate.categoria.slug
    return "combate"


def _nombre_categoria_registro(combate):
    detalle = _detalle_registro(combate)
    if detalle.get("nombre_categoria"):
        return detalle["nombre_categoria"]
    if combate.categoria and combate.categoria.nombre:
        return combate.categoria.nombre
    return "Combate"


def _ranking_figuras(combate):
    detalle = _detalle_registro(combate)
    ranking = detalle.get("ranking")
    return ranking if isinstance(ranking, list) else []


def _descripcion_registro(combate):
    """Descripción pública del registro (figuras): 'Intermedios 15-17 años'."""
    return (_detalle_registro(combate).get("descripcion") or "").strip()


def _llave_registro(combate):
    """Info de llave del registro (combates de eliminación) o None."""
    llave = _detalle_registro(combate).get("llave")
    return llave if isinstance(llave, dict) else None


def _es_final_llave(combate):
    """True si este registro es el combate final de una llave de eliminación."""
    llave = _llave_registro(combate)
    return bool(llave and (llave.get("ronda_nombre") or "").strip().lower() == "final")


def _podios_llaves_map(rows):
    """
    {llave_id: {nombre, podio}} para las llaves de eliminación referenciadas por
    las filas dadas que ya tienen campeón. Una sola consulta por lote.
    """
    ids = set()
    nombres = {}
    for combate, _t, _c in rows:
        llave = _llave_registro(combate)
        if llave and llave.get("llave_id"):
            ids.add(int(llave["llave_id"]))
            nombres[int(llave["llave_id"])] = llave.get("nombre") or combate.nombre_hong
    if not ids:
        return {}
    mapa = {}
    for ll in Llave.query.filter(Llave.id.in_(ids)).all():
        podio = podio_llave(ll.estructura)
        if podio:
            mapa[ll.id] = {"nombre": ll.nombre or nombres.get(ll.id, "-"), "podio": podio}
    return mapa


def _puesto_figuras_txt(item, t=None):
    """Puesto anotado: '1 (Especial)' / '3 (Empate)' para Excel y PDF."""
    t = t or _T_ES
    puesto = str(item.get("puesto", "-"))
    if item.get("especial"):
        puesto += t("especial_sfx")
    elif item.get("empate"):
        puesto += t("empate_sfx")
    return puesto


def _medalla_txt(puesto, t=None):
    t = t or _T_ES
    claves = {1: "medalla_oro", 2: "medalla_plata", 3: "medalla_bronce"}
    if puesto in claves:
        return t(claves[puesto])
    return t("medalla_n", n=puesto)


def _figuras_completas(combate):
    """True si todos los competidores fueron calificados en todos sus criterios."""
    detalle = _detalle_registro(combate)
    if "puntuaciones_completas" in detalle:
        return bool(detalle["puntuaciones_completas"])
    # Registros antiguos: derivar de confirmaciones
    comps = detalle.get("competidores") or []
    confirmadas = detalle.get("puntuaciones_confirmadas") or {}
    criterios = detalle.get("criterios") or []
    num_jueces = combate.num_jueces or 4
    jueces = [f"j{i}" for i in range(1, num_jueces + 1)][:len(criterios) or num_jueces]
    if not comps or not jueces:
        return False
    return all(
        confirmadas.get(str(c.get("id")), {}).get(j)
        for c in comps for j in jueces
    )


def _figuras_desempates(combate):
    """Reevaluaciones por empate registradas en la categoría."""
    desempates = _detalle_registro(combate).get("desempates")
    return desempates if isinstance(desempates, list) else []


def _figuras_estado(combate, t=None):
    t = t or _T_ES
    estado = t("completo") if _figuras_completas(combate) else t("incompleto")
    desempates = _figuras_desempates(combate)
    if desempates:
        nombres = "; ".join(
            " y ".join(d.get("nombres", [])) for d in desempates
        )
        estado += " " + t("desempate_reev", nombres=nombres)
    return estado


def _puntos_por_ronda(combate):
    """
    Puntos por ronda y color EN LA MISMA ESCALA que el marcador final:
    los puntos de jueces de esquina se promedian entre los jueces activos
    (igual que calcular_marcador) y los del árbitro van a valor completo.
    Así la suma de las rondas coincide con el total del combate.
    """
    num_jueces = combate.num_jueces or 4
    rondas = {}
    for entrada in (combate.historial_completo or []):
        color = entrada.get("color")
        if color not in ("hong", "chung"):
            continue
        try:
            pts = float(entrada.get("pts") or 0)
        except (TypeError, ValueError):
            pts = 0.0
        juez = str(entrada.get("juez") or "")
        if juez.startswith("j"):
            try:
                if int(juez[1:]) > num_jueces:
                    continue  # juez fuera de la configuración final del combate
            except ValueError:
                continue
            pts = pts / num_jueces
        ronda = entrada.get("ronda") or "-"
        slot = rondas.setdefault(ronda, {"hong": 0.0, "chung": 0.0})
        slot[color] += pts
    return rondas


def _fmt_pts(valor):
    return f"{round(valor, 2):g}"


def _rondas_resumen(combate, t=None):
    """Texto 'Round 1: 5-3 · Round 2: 2-2 · Punto de Oro: 1-0' segun historial."""
    t = t or _T_ES
    if _tipo_registro(combate) == "figuras":
        return t("figuras_estado", estado=_figuras_estado(combate, t))
    rondas = _puntos_por_ronda(combate)
    partes = []
    orden = ["r1", "r2", "oro"]
    for rid in orden + [r for r in rondas if r not in orden]:
        if rid in rondas:
            r = rondas[rid]
            etiqueta = t(RONDA_CLAVE[rid]) if rid in RONDA_CLAVE else rid
            partes.append(f"{etiqueta}: {_fmt_pts(r['hong'])}-{_fmt_pts(r['chung'])}")
    return " · ".join(partes) if partes else "-"


def _ronda_label(ronda, t=None):
    t = t or _T_ES
    if ronda in RONDA_CLAVE:
        return t(RONDA_CLAVE[ronda])
    return ronda or "-"


def _tipo_punto_txt(entrada, t=None):
    """Etiqueta traducida del tipo de punto de una entrada del historial."""
    t = t or _T_ES
    if entrada.get("esEspecial"):
        return t("pt_especial")
    if entrada.get("esKyongGo"):
        return t("pt_kyonggo")
    if entrada.get("esGamJeum"):
        return t("pt_gamjeum")
    if entrada.get("esDecision"):
        return t("pt_decision")
    return t("pt_normal")


def _figuras_puntajes_detalle(combate, comp_id):
    """Puntajes de un competidor con criterio, juez y correo."""
    detalle = _detalle_registro(combate)
    criterios = detalle.get("criterios") or []
    asignaciones = _jueces_meta(combate)
    puntajes = (detalle.get("puntuaciones") or {}).get(str(comp_id), {})
    partes = []
    for juez_id, valor in sorted(puntajes.items()):
        sufijo = juez_id[1:] if isinstance(juez_id, str) else ""
        idx = int(sufijo) - 1 if sufijo.isdigit() else None
        criterio = (
            criterios[idx]["nombre"]
            if idx is not None and 0 <= idx < len(criterios)
            else str(juez_id)
        )
        meta = asignaciones.get(juez_id, {})
        quien = meta.get("nombre") or str(juez_id).upper()
        email = meta.get("email") or ""
        etiqueta = f"{str(juez_id).upper()} {criterio} — {quien}"
        if email:
            etiqueta += f" <{email}>"
        try:
            partes.append(f"{etiqueta}: {float(valor):.2f}")
        except (TypeError, ValueError):
            partes.append(f"{etiqueta}: {valor}")
    return "; ".join(partes) or "-"


def _ganador_nombre(combate, t=None):
    t = t or _T_ES
    if _tipo_registro(combate) == "figuras":
        ranking = _ranking_figuras(combate)
        # 1° del podio normal (los especiales tienen su podio aparte)
        normal = next((r for r in ranking if not r.get("especial")), None)
        item = normal or (ranking[0] if ranking else None)
        return item.get("nombre", "-") if item else "-"
    if combate.ganador == "hong":
        return combate.nombre_hong
    if combate.ganador == "chung":
        return combate.nombre_chung
    return t("empate")


@reportes_bp.route("/combates", methods=["GET"])
@jwt_required()
def listar_combates():
    """GET /api/reportes/combates — Lista combates guardados con filtros."""
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    # Coerción defensiva: paginación con basura no debe responder 500
    try:
        page = max(1, int(request.args.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        per_page = min(200, max(1, int(request.args.get("per_page", 50))))
    except (TypeError, ValueError):
        per_page = 50

    # Todas las filas con filtros SQL: el resumen por categoría se calcula
    # ANTES del filtro de categoría para que el selector muestre todas.
    rows_sql = _build_query(request.args, user=admin).all()
    categorias = _resumen_categorias(rows_sql)

    categoria = (request.args.get("categoria") or "").strip().lower()
    rows_filtradas = rows_sql
    if categoria:
        rows_filtradas = [
            r for r in rows_sql
            if _nombre_categoria_registro(r[0]).strip().lower() == categoria
        ]

    total = len(rows_filtradas)
    rows = rows_filtradas[(page - 1) * per_page: page * per_page]

    podios_llaves = _podios_llaves_map(rows)
    t = trad(idioma_request())

    result = []
    for c, tatami, camp in rows:
        tipo = _tipo_registro(c)
        llave_info = _llave_registro(c)
        podio_combate = []
        if _es_final_llave(c) and llave_info and llave_info.get("llave_id"):
            podio_combate = podios_llaves.get(int(llave_info["llave_id"]), {}).get("podio", [])
        result.append({
            "id": c.id,
            "tipo": tipo,
            "nombre_categoria": _nombre_categoria_registro(c),
            "descripcion": _descripcion_registro(c),
            "nombre_hong": c.nombre_hong,
            "nombre_chung": c.nombre_chung,
            "marcador_hong": float(c.marcador_hong or 0),
            "marcador_chung": float(c.marcador_chung or 0),
            "ganador": c.ganador,
            "ronda_final": c.ronda_final,
            "rondas_resumen": _rondas_resumen(c, t),
            "figuras_completas": _figuras_completas(c) if tipo == "figuras" else None,
            "figuras_desempates": _figuras_desempates(c) if tipo == "figuras" else [],
            # Combate de eliminación: llave y ronda (Semifinal, Final, ...)
            "llave": (c.jueces_detalle or {}).get("llave"),
            "num_jueces": c.num_jueces,
            "duracion_segundos": c.duracion_segundos,
            "tatami_id": tatami.id if tatami else None,
            "tatami_numero": tatami.numero if tatami else None,
            "campeonato_id": camp.id if camp else None,
            "campeonato_nombre": camp.nombre if camp else None,
            # Con zona UTC explícita: el navegador los muestra en hora local
            "created_at": iso_utc(c.created_at),
            "fin": iso_utc(c.fin),
            "historial_completo": c.historial_completo or [],
            "jueces_detalle": c.jueces_detalle or {},
            "jueces": _jueces_list(c),
            "jueces_resumen": _jueces_resumen(c),
            "ranking": _ranking_figuras(c) if tipo == "figuras" else [],
            "podio_llave": podio_combate,
        })

    return jsonify({
        "combates": result,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
        "categorias": categorias,
    }), 200


# ══════════════════════════════════════════════════════════════════════════════
#  GENERADORES (reutilizados por export simple y export ZIP dividido)
# ══════════════════════════════════════════════════════════════════════════════

def _generar_excel(rows, subtitulo="", t=None):
    """Genera el workbook Excel para una lista de (combate, tatami, camp)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    t = t or _T_ES
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = t("hoja_resultados")

    # Estilos
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    hong_fill = PatternFill(start_color="FFE5E8", end_color="FFE5E8", fill_type="solid")
    chung_fill = PatternFill(start_color="E5EEFF", end_color="E5EEFF", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # Titulo
    ws.merge_cells("A1:P1")
    titulo = ws["A1"]
    titulo.value = f"{t('reporte_titulo')} — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    titulo.font = Font(name="Arial", bold=True, size=13, color="1A1A2E")
    titulo.alignment = center
    if subtitulo:
        ws.merge_cells("A2:P2")
        sub = ws["A2"]
        sub.value = subtitulo
        sub.font = Font(name="Arial", bold=True, size=11, color="555555")
        sub.alignment = center

    # Headers
    headers = [
        t("h_id"), t("h_tipo"), t("h_categoria"), t("h_campeonato"), t("h_tatami"),
        t("h_rojo_comp"), t("h_azul_cat"), t("h_pts1_total"), t("h_pts2"),
        t("h_ganador"), t("h_ronda_final"), t("h_rondas_hc"), t("h_num_jueces"),
        t("h_duracion"), t("h_fecha_hora"), t("h_jueces_full"),
    ]
    row = 3
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[row].height = 24

    # Datos
    for combate, tatami, camp in rows:
        tipo = _tipo_registro(combate)

        row += 1
        ganador_nombre = _ganador_nombre(combate, t)

        data = [
            combate.id,
            t("figuras") if tipo == "figuras" else t("combate"),
            _nombre_categoria_registro(combate),
            camp.nombre if camp else "-",
            t("tatami_n", n=tatami.numero) if tatami else "-",
            combate.nombre_hong,
            combate.nombre_chung,
            float(combate.marcador_hong or 0),
            float(combate.marcador_chung or 0),
            ganador_nombre,
            t("figuras") if tipo == "figuras" else _ronda_label(combate.ronda_final, t),
            _rondas_resumen(combate, t),
            combate.num_jueces or 4,
            combate.duracion_segundos or "-",
            _fecha_local_txt(combate.created_at),
            _jueces_resumen(combate),
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = center
            # Color by ganador
            if tipo != "figuras" and col_idx in (6, 8) and combate.ganador == "hong":
                cell.fill = hong_fill
            elif tipo != "figuras" and col_idx in (7, 9) and combate.ganador == "chung":
                cell.fill = chung_fill

    # Widths
    col_widths = [6, 12, 20, 22, 10, 22, 22, 12, 10, 20, 14, 30, 12, 12, 18, 48]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # ── Hoja 2: Detalle de puntos por juez ──────────────────────────────────
    ws2 = wb.create_sheet(t("hoja_detalle"))
    det_headers = [
        t("h_combate_id"), t("h_tatami"), t("h_hong"), t("h_chung"), t("h_rol"),
        t("h_nombre_juez"), t("h_correo"), t("h_asignacion"), t("h_acceso"),
        t("h_color"), t("h_pts"), t("h_accion"), t("h_tiempo"),
        t("h_momento"), t("h_ronda"), t("h_tipo"),
    ]
    for col_idx, h in enumerate(det_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    det_row = 2
    for combate, tatami, camp in rows:
        if _tipo_registro(combate) == "figuras":
            continue
        historial = combate.historial_completo or []
        for entrada in historial:
            juez_meta = _juez_meta_para_evento(combate, entrada)
            tipo = _tipo_punto_txt(entrada, t)

            det_data = [
                combate.id,
                t("tatami_n", n=tatami.numero) if tatami else "-",
                combate.nombre_hong,
                combate.nombre_chung,
                juez_meta["rol"],
                juez_meta["nombre"],
                juez_meta["email"],
                juez_meta["asignacion"],
                juez_meta["acceso"],
                entrada.get("color", "-"),
                entrada.get("pts", 0),
                entrada.get("nombre", "-"),
                entrada.get("tiempo", "-"),
                _format_momento_evento(entrada),
                _ronda_label(entrada.get("ronda"), t),
                tipo,
            ]
            for col_idx, val in enumerate(det_data, 1):
                cell = ws2.cell(row=det_row, column=col_idx, value=val)
                cell.border = border
                cell.alignment = center
                if entrada.get("color") == "hong":
                    cell.fill = hong_fill
                elif entrada.get("color") == "chung":
                    cell.fill = chung_fill
            det_row += 1

    for i, w in enumerate([10, 10, 20, 20, 10, 24, 30, 18, 12, 8, 8, 22, 10, 20, 14, 14], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Hoja 3: Ranking de Figuras ──────────────────────────────────────────
    ws3 = wb.create_sheet(t("hoja_ranking"))
    fig_headers = [
        t("h_registro_id"), t("h_categoria"), t("h_descripcion"), t("h_campeonato"),
        t("h_tatami"), t("h_estado"), t("h_puesto"),
        t("h_competidor"), t("h_club"), t("h_total"), t("h_puntajes_juez"),
    ]
    for col_idx, h in enumerate(fig_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    fig_row = 2
    for combate, tatami, camp in rows:
        if _tipo_registro(combate) != "figuras":
            continue
        estado = _figuras_estado(combate, t)
        for item in _ranking_figuras(combate):
            comp_id = str(item.get("id", ""))
            fig_data = [
                combate.id,
                _nombre_categoria_registro(combate),
                _descripcion_registro(combate) or "-",
                camp.nombre if camp else "-",
                t("tatami_n", n=tatami.numero) if tatami else "-",
                estado,
                _puesto_figuras_txt(item, t),
                item.get("nombre", "-"),
                item.get("club") or "-",
                float(item.get("total", 0)),
                _figuras_puntajes_detalle(combate, comp_id),
            ]
            for col_idx, val in enumerate(fig_data, 1):
                cell = ws3.cell(row=fig_row, column=col_idx, value=val)
                cell.border = border
                cell.alignment = center
            fig_row += 1

    for i, w in enumerate([12, 22, 24, 22, 10, 12, 8, 24, 18, 10, 64], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Hoja 4: Podios de Llaves (eliminación) ──────────────────────────────
    podios_map = _podios_llaves_map(rows)
    if podios_map:
        ws4 = wb.create_sheet(t("hoja_podios"))
        po_headers = [t("h_llave"), t("h_puesto"), t("h_competidor"), t("h_club")]
        for col_idx, h in enumerate(po_headers, 1):
            cell = ws4.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        po_row = 2
        for _llid, info in sorted(podios_map.items(), key=lambda x: x[1]["nombre"]):
            for item in info["podio"]:
                po_data = [
                    info["nombre"],
                    _medalla_txt(item["puesto"], t),
                    item.get("nombre", "-"),
                    item.get("club") or "-",
                ]
                for col_idx, val in enumerate(po_data, 1):
                    cell = ws4.cell(row=po_row, column=col_idx, value=val)
                    cell.border = border
                    cell.alignment = center
                po_row += 1
        for i, w in enumerate([28, 12, 26, 22], 1):
            ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Output
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _generar_pdf(rows, subtitulo="", t=None):
    """Genera el documento PDF para una lista de (combate, tatami, camp)."""
    t = t or _T_ES
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable
    )

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    story = []

    # Titulo
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#1A1A2E"),
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#666666"),
        spaceAfter=12,
    )

    story.append(Paragraph(t("reporte_titulo"), title_style))
    encabezado = (
        f"Global Hapkido Association &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"{t('reporte_generado', fecha=datetime.now().strftime('%d/%m/%Y %H:%M'))} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"{t('reporte_total', n=len(rows))}"
    )
    if subtitulo:
        encabezado = f"<b>{subtitulo}</b> &nbsp;&nbsp;|&nbsp;&nbsp; " + encabezado
    story.append(Paragraph(encabezado, sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#DDDDDD")))
    story.append(Spacer(1, 0.3 * cm))

    # Headers
    col_widths_pdf = [
        1*cm, 1.8*cm, 3.8*cm, 1.5*cm, 3.2*cm, 3.2*cm,
        1.5*cm, 1.5*cm, 3.0*cm, 1.8*cm, 2.4*cm
    ]

    table_data = [[t("h_num"), t("h_tipo"), t("h_campeonato"), t("h_tatami"),
                   t("h_rojo_comp_corto"), t("h_azul_cat_corto"),
                   t("h_pts1_total"), t("h_pts2"), t("h_ganador"), t("h_ronda"), t("h_fecha")]]

    for combate, tatami, camp in rows:
        tipo = _tipo_registro(combate)

        table_data.append([
            str(combate.id),
            t("fig_corto") if tipo == "figuras" else t("comb_corto"),
            camp.nombre if camp else "-",
            t("tatami_corto", n=tatami.numero) if tatami else "-",
            combate.nombre_hong or "-",
            combate.nombre_chung or "-",
            str(float(combate.marcador_hong or 0)),
            str(float(combate.marcador_chung or 0)),
            _ganador_nombre(combate, t),
            t("fig_corto") if tipo == "figuras" else _ronda_label(combate.ronda_final, t),
            _fecha_local_txt(combate.created_at, "%d/%m %H:%M"),
        ])

    # Colors
    DARK = colors.HexColor("#1A1A2E")
    HONG_BG = colors.HexColor("#FFE5E8")
    CHUNG_BG = colors.HexColor("#E5EEFF")
    GRAY_BG = colors.HexColor("#F5F5F5")
    WHITE = colors.white

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, GRAY_BG]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    # Highlight winner cells
    for row_idx, (combate, _t, _c) in enumerate(rows, 1):
        if _tipo_registro(combate) == "figuras":
            continue
        if combate.ganador == "hong":
            style_cmds.append(("BACKGROUND", (4, row_idx), (4, row_idx), HONG_BG))
            style_cmds.append(("BACKGROUND", (6, row_idx), (6, row_idx), HONG_BG))
        elif combate.ganador == "chung":
            style_cmds.append(("BACKGROUND", (5, row_idx), (5, row_idx), CHUNG_BG))
            style_cmds.append(("BACKGROUND", (7, row_idx), (7, row_idx), CHUNG_BG))

    tabla = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
    tabla.setStyle(TableStyle(style_cmds))
    story.append(tabla)

    # ── Marcador por rondas (R1 / R2 / Punto de Oro) ──
    combates_normales = [r for r in rows if _tipo_registro(r[0]) != "figuras"]
    if combates_normales:
        story.append(Spacer(1, 0.7 * cm))
        story.append(Paragraph(t("sec_marcador_rondas"), title_style))
        rondas_table_data = [[t("h_comb_id"), t("h_rojo"), t("h_azul"), t("h_rondas_jugadas")]]
        for combate, _t2, _c2 in combates_normales:
            rondas_table_data.append([
                str(combate.id),
                combate.nombre_hong or "-",
                combate.nombre_chung or "-",
                _rondas_resumen(combate, t),
            ])
        rondas_table = Table(
            rondas_table_data, colWidths=[2*cm, 5*cm, 5*cm, 14*cm], repeatRows=1
        )
        rondas_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, GRAY_BG]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(rondas_table)

    # ── Resumen de jueces por combate ──
    story.append(Spacer(1, 0.7 * cm))
    story.append(Paragraph(t("sec_jueces"), title_style))
    jueces_table_data = [[t("h_comb_id"), t("h_tatami"), t("h_jueces")]]
    for combate, tatami, _c3 in rows:
        jueces_table_data.append([
            str(combate.id),
            t("tatami_corto", n=tatami.numero) if tatami else "-",
            _jueces_resumen(combate),
        ])

    jueces_table = Table(jueces_table_data, colWidths=[2*cm, 1.6*cm, 22.4*cm], repeatRows=1)
    jueces_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(jueces_table)

    # ── Detalle de puntos en PDF ──
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph(t("sec_detalle_puntos"), title_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#DDDDDD")))
    story.append(Spacer(1, 0.5 * cm))

    det_col_widths = [1.3*cm, 1.2*cm, 3*cm, 4.2*cm, 2.4*cm, 1.4*cm, 1.2*cm, 1*cm, 2.6*cm, 2.6*cm, 1.7*cm, 1.8*cm]
    det_table_data = [[t("h_comb_id"), t("h_rol"), t("h_nombre"), t("h_correo"),
                       t("h_asign_corto"), t("h_acceso"), t("h_color"), t("h_pts"),
                       t("h_accion"), t("h_momento"), t("h_ronda"), t("h_tipo")]]

    det_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    det_row_idx = 1
    for combate, _t4, _c4 in rows:
        if _tipo_registro(combate) == "figuras":
            continue
        historial = combate.historial_completo or []
        for entrada in historial:
            juez_meta = _juez_meta_para_evento(combate, entrada)
            tipo = _tipo_punto_txt(entrada, t)

            det_table_data.append([
                str(combate.id),
                str(juez_meta["rol"]),
                str(juez_meta["nombre"]),
                str(juez_meta["email"]),
                str(juez_meta["asignacion"]),
                str(juez_meta["acceso"]),
                str(entrada.get("color", "-")),
                str(entrada.get("pts", 0)),
                str(entrada.get("nombre", "-")),
                _format_momento_evento(entrada),
                _ronda_label(entrada.get("ronda"), t),
                tipo,
            ])

            # Colores de fila alternos y colores para Hong/Chung
            bg_color = WHITE if det_row_idx % 2 != 0 else GRAY_BG
            det_style_cmds.append(("BACKGROUND", (0, det_row_idx), (-1, det_row_idx), bg_color))

            if entrada.get("color") == "hong":
                det_style_cmds.append(("BACKGROUND", (6, det_row_idx), (6, det_row_idx), HONG_BG))
            elif entrada.get("color") == "chung":
                det_style_cmds.append(("BACKGROUND", (6, det_row_idx), (6, det_row_idx), CHUNG_BG))

            det_row_idx += 1

    if len(det_table_data) > 1:
        t_det = Table(det_table_data, colWidths=det_col_widths, repeatRows=1)
        t_det.setStyle(TableStyle(det_style_cmds))
        story.append(t_det)

    # ── Ranking de Figuras en PDF ───────────────────────────────────────────
    figuras = [r for r in rows if _tipo_registro(r[0]) == "figuras"]
    if figuras:
        story.append(Spacer(1, 1 * cm))
        story.append(Paragraph(t("sec_ranking_figuras"), title_style))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#DDDDDD")))
        story.append(Spacer(1, 0.4 * cm))

        fig_table_data = [[t("h_reg_id"), t("h_categoria"), t("h_descripcion"),
                           t("h_estado"), t("h_puesto"), t("h_competidor"), t("h_club"), t("h_total")]]
        for combate, _t5, _c5 in figuras:
            estado = _figuras_estado(combate, t)
            descripcion = _descripcion_registro(combate) or "-"
            for item in _ranking_figuras(combate):
                fig_table_data.append([
                    str(combate.id),
                    _nombre_categoria_registro(combate),
                    descripcion,
                    estado,
                    _puesto_figuras_txt(item, t),
                    str(item.get("nombre", "-")),
                    str(item.get("club") or "-"),
                    str(float(item.get("total", 0))),
                ])

        fig_table = Table(
            fig_table_data,
            colWidths=[1.4*cm, 3.8*cm, 3.4*cm, 2.0*cm, 1.3*cm, 4.8*cm, 3.4*cm, 1.5*cm],
            repeatRows=1,
        )
        fig_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, GRAY_BG]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ]))
        story.append(fig_table)

        # Detalle de puntajes por juez (criterio + correo)
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph(t("sec_puntajes_figuras"), title_style))
        figdet_data = [[t("h_reg_id"), t("h_competidor"), t("h_puntajes")]]
        for combate, _t6, _c6 in figuras:
            for item in _ranking_figuras(combate):
                figdet_data.append([
                    str(combate.id),
                    str(item.get("nombre", "-")),
                    _figuras_puntajes_detalle(combate, str(item.get("id", ""))),
                ])
        figdet_table = Table(figdet_data, colWidths=[1.6*cm, 5*cm, 19.4*cm], repeatRows=1)
        figdet_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, GRAY_BG]),
        ]))
        story.append(figdet_table)

    # ── Podios de Llaves (eliminación) en PDF ───────────────────────────────
    podios_map = _podios_llaves_map(rows)
    if podios_map:
        story.append(Spacer(1, 1 * cm))
        story.append(Paragraph(t("sec_podios_llaves"), title_style))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#DDDDDD")))
        story.append(Spacer(1, 0.4 * cm))
        po_data = [[t("h_llave"), t("h_puesto"), t("h_competidor"), t("h_club")]]
        for _llid, info in sorted(podios_map.items(), key=lambda x: x[1]["nombre"]):
            for item in info["podio"]:
                po_data.append([
                    info["nombre"],
                    _medalla_txt(item["puesto"], t),
                    str(item.get("nombre", "-")),
                    str(item.get("club") or "-"),
                ])
        po_table = Table(po_data, colWidths=[7.5*cm, 2.6*cm, 8.5*cm, 7.4*cm], repeatRows=1)
        po_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, GRAY_BG]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ]))
        story.append(po_table)

    doc.build(story)
    output.seek(0)
    return output


# ══════════════════════════════════════════════════════════════════════════════
#  ENDPOINTS DE EXPORTACIÓN
# ══════════════════════════════════════════════════════════════════════════════

@reportes_bp.route("/combates/export/excel", methods=["GET"])
@jwt_required()
def exportar_excel():
    """GET /api/reportes/combates/export/excel — Exportar combates en Excel."""
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    try:
        t = trad(idioma_request())
        rows = _rows_filtradas(request.args, user=admin)
        ctx = _contexto_filtros(request.args, t)
        output = _generar_excel(rows, _subtitulo_filtros(ctx, t), t)
    except ImportError:
        return jsonify({
            "error": "No se pudo generar Excel: falta openpyxl. Instala dependencias con pip install -r requirements.txt."
        }), 500

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=_nombre_archivo(ctx, "xlsx"),
    )


@reportes_bp.route("/combates/export/pdf", methods=["GET"])
@jwt_required()
def exportar_pdf():
    """GET /api/reportes/combates/export/pdf — Exportar combates en PDF."""
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    try:
        t = trad(idioma_request())
        rows = _rows_filtradas(request.args, user=admin)
        ctx = _contexto_filtros(request.args, t)
        output = _generar_pdf(rows, _subtitulo_filtros(ctx, t), t)
    except ImportError:
        return jsonify({
            "error": "No se pudo generar PDF: falta reportlab. Instala dependencias con pip install -r requirements.txt."
        }), 500

    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=_nombre_archivo(ctx, "pdf"),
    )


@reportes_bp.route("/combates/export/zip", methods=["GET"])
@jwt_required()
def exportar_zip():
    """
    GET /api/reportes/combates/export/zip?formato=excel|pdf
    Exporta los reportes divididos en un ZIP:
    - Con tatami_id: un archivo con solo ese tatami.
    - Con campeonato_id: un archivo por cada tatami del campeonato.
    - Sin filtros: un archivo por cada campeonato.
    """
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    formato = request.args.get("formato", "excel")
    if formato not in ("excel", "pdf"):
        return jsonify({"error": "Formato inválido. Usa 'excel' o 'pdf'."}), 400

    rows = _rows_filtradas(request.args, user=admin)
    if not rows:
        return jsonify({"error": "No hay registros con los filtros actuales."}), 404

    t = trad(idioma_request())
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    ext = "xlsx" if formato == "excel" else "pdf"
    generar = _generar_excel if formato == "excel" else _generar_pdf

    # Agrupar: por tatami dentro de un campeonato, o por campeonato global
    grupos = []  # (nombre_archivo, subtitulo, rows_grupo)
    if request.args.get("campeonato_id") or request.args.get("tatami_id"):
        por_tatami = {}
        for row in rows:
            _combate, tatami, camp = row
            clave = tatami.id if tatami else 0
            por_tatami.setdefault(clave, []).append(row)
        for clave in sorted(por_tatami):
            grupo = por_tatami[clave]
            _cb, tatami, camp = grupo[0]
            camp_nombre = camp.nombre if camp else "-"
            tatami_label = t("tatami_n", n=tatami.numero) if tatami else "-"
            nombre = f"dinamyt_{_slug(camp_nombre, 'campeonato')}_tatami{tatami.numero if tatami else 0}_{fecha}.{ext}"
            subtitulo = f"{t('sub_campeonato', v=camp_nombre)} — {tatami_label}"
            grupos.append((nombre, subtitulo, grupo))
        zip_nombre = f"dinamyt_{_slug(grupos[0][2][0][2].nombre if grupos[0][2][0][2] else 'campeonato', 'campeonato')}_por_tatami_{fecha}.zip"
    else:
        por_camp = {}
        for row in rows:
            _combate, tatami, camp = row
            clave = camp.id if camp else 0
            por_camp.setdefault(clave, []).append(row)
        for clave in sorted(por_camp):
            grupo = por_camp[clave]
            _cb, _tt, camp = grupo[0]
            camp_nombre = camp.nombre if camp else "-"
            nombre = f"dinamyt_{_slug(camp_nombre, 'campeonato')}_{fecha}.{ext}"
            grupos.append((nombre, t("sub_campeonato", v=camp_nombre), grupo))
        zip_nombre = f"dinamyt_reportes_por_campeonato_{fecha}.zip"

    try:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            usados = set()
            for nombre, subtitulo, grupo in grupos:
                # Evitar nombres duplicados dentro del zip
                base = nombre
                contador = 2
                while nombre in usados:
                    nombre = base.replace(f".{ext}", f"_{contador}.{ext}")
                    contador += 1
                usados.add(nombre)
                archivo = generar(grupo, subtitulo, t)
                zf.writestr(nombre, archivo.getvalue())
        zip_buffer.seek(0)
    except ImportError:
        libreria = "openpyxl" if formato == "excel" else "reportlab"
        return jsonify({
            "error": f"No se pudo generar el ZIP: falta {libreria}. Instala dependencias con pip install -r requirements.txt."
        }), 500

    return send_file(
        zip_buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=zip_nombre,
    )
