"""
API: Competidores e Inscripciones

- Competidores: atletas registrados en el sistema (independientes del
  campeonato). Se pueden crear uno a uno o importar desde Excel.
- Inscripciones: vínculo competidor ↔ campeonato con modalidades y snapshot
  de peso/cinturón. Un competidor se inscribe desde su ficha o directamente
  al importar el Excel de un campeonato.
"""

import io
import re
import unicodedata
from datetime import datetime, date

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from sqlalchemy import func

from ..extensions import db
from ..security import limitar
from ..filei18n import trad, idioma_request
from ..models.campeonato import Campeonato
from ..models.competidor import (
    CINTURONES,
    Competidor,
    Inscripcion,
    cinturones_localizados,
    normalizar_cinturon,
)
from .scoping import (
    es_dueno_campeonato,
    es_dueno_competidor,
    filtrar_campeonatos,
    filtrar_competidores,
    require_admin as _require_admin,
    require_maestro,
    usuario_actual,
    workspace_owner_id,
)

competidores_bp = Blueprint("competidores", __name__)
inscripciones_bp = Blueprint("inscripciones", __name__)

MAX_IMPORT_FILAS = 1000

# Límites de los campos del competidor (validados también en el frontend)
NOMBRE_MIN = 8
NOMBRE_MAX = 80
DOCUMENTO_MAX = 15
CLUB_MAX = 80
EDAD_MIN, EDAD_MAX = 3, 100
PESO_MIN, PESO_MAX = 10.0, 200.0
PESO_CHARS_MAX = 6  # máx. caracteres en el string de peso (ej: "200.50")


def _mayusculas(valor):
    """Un NOMBRE (de persona o de club), tal y como se guarda: en MAYÚSCULAS.

    Misma regla —y mismo motivo— que `mayusculas` en `api/auth.py`: la planilla
    de inscritos, la llave impresa y el acta de resultados tienen que decir lo
    mismo aunque el nombre lo haya escrito el maestro, el admin o una hoja de
    Excel importada. `str.upper()` conserva tildes y eñes.
    """
    return None if valor is None else str(valor).upper()


def _sin_acentos(texto):
    return "".join(
        c for c in unicodedata.normalize("NFD", str(texto))
        if unicodedata.category(c) != "Mn"
    )


def _norm_genero(valor):
    """Normaliza el género (español o inglés) → MASCULINO | FEMENINO."""
    v = _sin_acentos(valor or "").strip().upper()
    if not v:
        return None
    # FEM cubre FEMENINO/FEMALE; añadimos MUJER y WOMAN
    if v == "F" or v.startswith(("FEM", "MUJER", "WOMAN")):
        return "FEMENINO"
    # MASC cubre MASCULINO; añadimos inglés MALE/MAN
    if v in ("M", "H") or v.startswith(("MASC", "HOM", "VARON", "MALE", "MAN")):
        return "MASCULINO"
    return None


def _parse_fecha(valor):
    """Acepta date/datetime o texto ISO, dd/mm/aaaa y dd-mm-aaaa."""
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _validar_fecha_nacimiento(valor):
    """Retorna (fecha, error): la edad debe estar entre 3 y 100 años."""
    fecha = _parse_fecha(valor)
    if fecha is None:
        return None, "Fecha de nacimiento inválida (usa dd/mm/aaaa)."
    hoy = date.today()
    edad = hoy.year - fecha.year - (
        (hoy.month, hoy.day) < (fecha.month, fecha.day)
    )
    if edad < EDAD_MIN:
        return None, f"La edad mínima es {EDAD_MIN} años."
    if edad > EDAD_MAX:
        return None, f"Fecha de nacimiento errónea: da más de {EDAD_MAX} años."
    return fecha, None


def _validar_peso(valor):
    """Retorna (peso, error): entre 10 y 200 kg, coma o punto decimal."""
    try:
        texto_peso = str(valor).replace(",", ".")
        if len(texto_peso) > PESO_CHARS_MAX:
            return None, f"El peso no puede tener más de {PESO_CHARS_MAX} caracteres."
        peso = float(texto_peso)
    except (TypeError, ValueError):
        return None, "Peso inválido: escribe un número en kilogramos."
    if not (PESO_MIN <= peso <= PESO_MAX):
        return None, f"El peso debe estar entre {PESO_MIN:g} y {PESO_MAX:g} kg."
    return round(peso, 2), None


def _validar_documento(valor):
    """Solo dígitos (se toleran espacios, puntos y guiones al escribirlo)."""
    limpio = re.sub(r"[.\-\s]", "", str(valor or "").strip())
    if not limpio:
        return None, None
    if not limpio.isdigit():
        return None, "El documento solo puede contener números."
    if len(limpio) > DOCUMENTO_MAX:
        return None, f"El documento no puede superar {DOCUMENTO_MAX} dígitos."
    return limpio, None


def _parse_bool(valor):
    return str(valor).strip().lower() in ("1", "true", "si", "sí", "s", "x", "yes")


def _club_del_maestro(maestro, pedido):
    """(club, error): a qué dojang del maestro se apunta el alumno.

    Un maestro puede dirigir varios (ver `clubes` en models/usuario.py), así que
    ya no vale imponer uno: el alumno de un dojang no puede acabar compitiendo a
    nombre de otro solo porque estaba primero en la lista. Pero tampoco se toma
    lo que llegue del cliente tal cual — solo se acepta un club QUE SEA SUYO;
    si no, cualquier maestro inscribiría gente a nombre de un club ajeno.

    Sin `club` en el cuerpo se usa el principal, que es lo que hacía siempre y
    lo que necesita un maestro con un solo dojang.
    """
    disponibles = maestro.clubes
    if not disponibles:
        return None, "Tu administrador aún no te asignó un club."

    elegido = str(pedido or "").strip()
    if not elegido:
        return disponibles[0], None
    if not maestro.dirige_club(elegido):
        return None, f"El club '{elegido}' no es uno de los tuyos."
    # Se devuelve el nombre tal y como está guardado, no como vino escrito: así
    # el club del alumno coincide letra por letra con el del maestro y las
    # agrupaciones por club de los reportes no se parten en dos.
    return next(c for c in disponibles if c.casefold() == elegido.casefold()), None


def _aplicar_datos(comp, data, parciales=False, actor=None):
    """
    Aplica los campos del body a un competidor validándolos. Con
    `parciales=True` (import) solo toca los campos que traen valor.
    El grupo de cinturón NO se recibe: se deriva del cinturón elegido.
    `actor` (admin actual) decide cuánto detalle mostrar si el documento
    choca con un competidor de otro workspace.
    Retorna un mensaje de error o None.
    """
    if "nombre_completo" in data or not parciales:
        nombre = str(data.get("nombre_completo") or "").strip()
        if not nombre and not parciales:
            return "El nombre del competidor es requerido"
        if nombre and len(nombre) < NOMBRE_MIN:
            return f"El nombre debe tener al menos {NOMBRE_MIN} caracteres (nombre completo)."
        if len(nombre) > NOMBRE_MAX:
            return f"El nombre no puede superar {NOMBRE_MAX} caracteres."
        if nombre:
            comp.nombre_completo = _mayusculas(nombre)

    if "documento" in data:
        crudo = str(data.get("documento") or "").strip()
        doc, error = _validar_documento(crudo)
        if error:
            return error
        if doc and doc != comp.documento:
            otro = Competidor.query.filter_by(documento=doc).first()
            if otro and otro.id != comp.id:
                # El documento es único en TODO el sistema. Si el homónimo es
                # de otro workspace, no revelar su nombre.
                if actor is None or es_dueno_competidor(actor, otro):
                    quien = f" ({otro.nombre_completo})"
                else:
                    quien = " (registrado por otro administrador)"
                return f"Ya existe un competidor con documento {doc}{quien}"
        if doc or not parciales:
            comp.documento = doc

    if "fecha_nacimiento" in data:
        crudo = data.get("fecha_nacimiento")
        if crudo in (None, ""):
            if not parciales:
                comp.fecha_nacimiento = None
        else:
            fecha, error = _validar_fecha_nacimiento(crudo)
            if error:
                return error
            comp.fecha_nacimiento = fecha

    if "genero" in data:
        crudo = data.get("genero")
        genero = _norm_genero(crudo)
        if crudo not in (None, "") and genero is None:
            return "Género inválido: usa M o F."
        if genero or not parciales:
            comp.genero = genero

    # El cinturón viene del catálogo y fija el grupo automáticamente.
    if "cinturon" in data:
        crudo = str(data.get("cinturon") or "").strip()
        if not crudo:
            if not parciales:
                comp.cinturon = None
                comp.grupo_cinturon = None
        else:
            nombre_cint, grupo = normalizar_cinturon(crudo)
            if not nombre_cint:
                validos = ", ".join(n for n, _ in CINTURONES)
                return f"Cinturón '{crudo}' no reconocido. Válidos: {validos}."
            comp.cinturon = nombre_cint
            comp.grupo_cinturon = grupo

    if "peso" in data:
        crudo = data.get("peso")
        if crudo in (None, ""):
            if not parciales:
                comp.peso = None
        else:
            peso, error = _validar_peso(crudo)
            if error:
                return error
            comp.peso = peso

    if "club" in data:
        club = str(data.get("club") or "").strip() or None
        if club and len(club) > CLUB_MAX:
            return f"El club no puede superar {CLUB_MAX} caracteres."
        if club or not parciales:
            comp.club = _mayusculas(club)

    if "categoria_especial" in data:
        crudo = data.get("categoria_especial")
        if isinstance(crudo, bool):
            comp.categoria_especial = crudo
        elif crudo not in (None, ""):
            comp.categoria_especial = _parse_bool(crudo)
        elif not parciales:
            comp.categoria_especial = False

    if "activo" in data and not parciales:
        comp.activo = bool(data.get("activo"))

    return None


# ══════════════════════════════════════════════════════════════════
#  CRUD de competidores
# ══════════════════════════════════════════════════════════════════

@competidores_bp.route("", methods=["GET"])
@jwt_required()
def listar():
    """
    GET /api/competidores?q=&include_inactivos=1
    Lista competidores con su número de inscripciones.
    """
    q = (request.args.get("q") or "").strip()
    include_inactivos = request.args.get("include_inactivos") in ("1", "true")

    # Workspace: un admin solo ve los competidores que él registró.
    user = usuario_actual()
    query = Competidor.query
    if user and user.rol == "admin":
        query = filtrar_competidores(user, query)
    if not include_inactivos:
        query = query.filter_by(activo=True)
    if q:
        patron = f"%{q}%"
        query = query.filter(
            db.or_(
                Competidor.nombre_completo.ilike(patron),
                Competidor.documento.ilike(patron),
                Competidor.club.ilike(patron),
            )
        )
    comps = query.order_by(Competidor.nombre_completo.asc()).limit(2000).all()

    conteos = dict(
        db.session.query(Inscripcion.competidor_id, func.count(Inscripcion.id))
        .group_by(Inscripcion.competidor_id)
        .all()
    )
    return jsonify([
        c.to_dict(num_inscripciones=conteos.get(c.id, 0)) for c in comps
    ]), 200


@competidores_bp.route("", methods=["POST"])
@jwt_required()
def crear():
    """POST /api/competidores — Crear un competidor en el sistema."""
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    data = request.get_json() or {}
    comp = Competidor(nombre_completo="", activo=True, created_by=admin.id)
    error = _aplicar_datos(comp, data, actor=admin)
    if error:
        return jsonify({"error": error}), 400
    db.session.add(comp)
    db.session.commit()
    return jsonify({
        "message": f"Competidor '{comp.nombre_completo}' registrado",
        "competidor": comp.to_dict(num_inscripciones=0),
    }), 201


@competidores_bp.route("/<int:comp_id>", methods=["PUT"])
@jwt_required()
def editar(comp_id):
    """PUT /api/competidores/:id — Editar datos de un competidor."""
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    comp = Competidor.query.get_or_404(comp_id)
    if not es_dueno_competidor(admin, comp):
        return jsonify({"error": "Competidor no encontrado"}), 404
    data = request.get_json() or {}
    error = _aplicar_datos(comp, data, actor=admin)
    if error:
        return jsonify({"error": error}), 400
    db.session.commit()
    return jsonify({
        "message": "Competidor actualizado",
        "competidor": comp.to_dict(),
    }), 200


@competidores_bp.route("/<int:comp_id>", methods=["DELETE"])
@jwt_required()
def eliminar(comp_id):
    """
    DELETE /api/competidores/:id — Eliminar un competidor del sistema
    (borra también sus inscripciones; las llaves ya generadas no cambian).
    """
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    comp = Competidor.query.get_or_404(comp_id)
    if not es_dueno_competidor(admin, comp):
        return jsonify({"error": "Competidor no encontrado"}), 404
    nombre = comp.nombre_completo
    db.session.delete(comp)
    db.session.commit()
    return jsonify({"message": f"Competidor '{nombre}' eliminado"}), 200


# ══════════════════════════════════════════════════════════════════
#  Plantilla e importación desde Excel
# ══════════════════════════════════════════════════════════════════

# Encabezado normalizado → campo interno. El grupo de cinturón ya no se
# importa: se deriva del cinturón.
COLUMNAS_EXCEL = {
    # Español
    "nombre": "nombre_completo",
    "nombre completo": "nombre_completo",
    "documento": "documento",
    "identificacion": "documento",
    "cedula": "documento",
    "fecha nacimiento": "fecha_nacimiento",
    "fecha de nacimiento": "fecha_nacimiento",
    "nacimiento": "fecha_nacimiento",
    "genero": "genero",
    "sexo": "genero",
    "cinturon": "cinturon",
    "peso": "peso",
    "peso kg": "peso",
    "club": "club",
    "academia": "club",
    "equipo": "club",
    "especial": "categoria_especial",
    "categoria especial": "categoria_especial",
    "modalidades": "modalidades",
    # Inglés (para importar listas en inglés)
    "name": "nombre_completo",
    "full name": "nombre_completo",
    "id": "documento",
    "id number": "documento",
    "document": "documento",
    "birth date": "fecha_nacimiento",
    "date of birth": "fecha_nacimiento",
    "birth": "fecha_nacimiento",
    "dob": "fecha_nacimiento",
    "gender": "genero",
    "sex": "genero",
    "belt": "cinturon",
    "weight": "peso",
    "weight kg": "peso",
    "academy": "club",
    "team": "club",
    "special": "categoria_especial",
    "special category": "categoria_especial",
    "disciplines": "modalidades",
    "modalities": "modalidades",
}


@competidores_bp.route("/plantilla", methods=["GET"])
@jwt_required()
def plantilla_excel():
    """GET /api/competidores/plantilla?lang=en|es — Plantilla .xlsx para importar."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    lang = idioma_request()
    t = trad(lang)
    cinturones = cinturones_localizados(lang)
    especial_si = "YES" if lang == "en" else "SI"
    especial_no = "NO"
    ejemplo_especial = especial_no
    ejemplo_cinturon = "Blue" if lang == "en" else "Azul"
    ejemplo_modalidades = "COMBAT, OPEN-HAND FORM" if lang == "en" else "COMBATE, FIGURA A MANOS LIBRES"

    wb = Workbook()
    ws = wb.active
    ws.title = t("tpl_hoja")
    headers = [
        t("tpl_h_nombre"), t("tpl_h_documento"), t("tpl_h_fecha"), t("tpl_h_genero"),
        t("tpl_h_cinturon"), t("tpl_h_peso"), t("tpl_h_club"), t("tpl_h_especial"),
        t("tpl_h_modalidades"),
    ]
    ws.append(headers)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1A1A2E")
    for celda in ws[1]:
        celda.font = bold
        celda.fill = fill
    ws.append([
        "Ana María Pérez", "1002003001", "15/04/2010", "F",
        ejemplo_cinturon, 48.5, "Club Dinamyt", ejemplo_especial, ejemplo_modalidades,
    ])
    anchos = [28, 14, 16, 10, 16, 8, 20, 10, 34]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

    # Desplegables en la hoja: género, cinturón (catálogo) y especial.
    filas_dv = 500
    dv_genero = DataValidation(type="list", formula1='"M,F"', allow_blank=True)
    nombres_cinturones = ",".join(cinturones)
    dv_cinturon = DataValidation(
        type="list", formula1=f'"{nombres_cinturones}"', allow_blank=True
    )
    dv_especial = DataValidation(
        type="list", formula1=f'"{especial_si},{especial_no}"', allow_blank=True
    )
    ws.add_data_validation(dv_genero)
    ws.add_data_validation(dv_cinturon)
    ws.add_data_validation(dv_especial)
    dv_genero.add(f"D2:D{filas_dv}")
    dv_cinturon.add(f"E2:E{filas_dv}")
    dv_especial.add(f"H2:H{filas_dv}")

    notas = wb.create_sheet(t("tpl_hoja_notas"))
    notas["A1"] = t("tpl_notas_titulo")
    notas["A1"].font = Font(bold=True, size=13)
    filas_notas = [
        t("tpl_n_nombre", n=NOMBRE_MAX),
        t("tpl_n_documento", n=DOCUMENTO_MAX),
        t("tpl_n_fecha", min=EDAD_MIN, max=EDAD_MAX),
        t("tpl_n_genero"),
        t("tpl_n_cinturon", lista=", ".join(cinturones)),
        t("tpl_n_peso", min=f"{PESO_MIN:g}", max=f"{PESO_MAX:g}"),
        t("tpl_n_club", n=CLUB_MAX),
        t("tpl_n_especial"),
        t("tpl_n_modalidades"),
        t("tpl_n_ejemplo"),
    ]
    for i, texto in enumerate(filas_notas, start=3):
        notas[f"A{i}"] = texto
    notas.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = ("dinamyt_competitors_template.xlsx" if lang == "en"
             else "dinamyt_plantilla_competidores.xlsx")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


def _leer_excel(archivo):
    """
    Lee el .xlsx y retorna (filas, error). Cada fila es un dict campo→valor
    según los encabezados reconocidos de la primera fila.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
    except Exception:
        return None, "No se pudo leer el archivo. Debe ser un .xlsx válido."

    ws = wb.active
    filas = ws.iter_rows(values_only=True)
    try:
        encabezados = next(filas)
    except StopIteration:
        return None, "El archivo está vacío."

    mapa = {}
    for idx, enc in enumerate(encabezados or []):
        clave = _sin_acentos(enc or "").strip().lower()
        campo = COLUMNAS_EXCEL.get(clave)
        if campo and campo not in mapa.values():
            mapa[idx] = campo
    if "nombre_completo" not in mapa.values():
        return None, (
            "No se encontró la columna de nombre. Usa la plantilla o incluye "
            "un encabezado 'Nombre completo'."
        )

    resultado = []
    for numero, fila in enumerate(filas, start=2):
        if fila is None:
            continue
        datos = {"_fila": numero}
        for idx, campo in mapa.items():
            valor = fila[idx] if idx < len(fila) else None
            if valor is not None and str(valor).strip() != "":
                datos[campo] = valor
        if len(datos) > 1:
            resultado.append(datos)
        if len(resultado) > MAX_IMPORT_FILAS:
            return None, f"Máximo {MAX_IMPORT_FILAS} filas por importación."
    wb.close()
    return resultado, None


def _parse_modalidades(valor):
    """'COMBATE, FIGURA CON ARMAS' → ['COMBATE', 'FIGURA CON ARMAS']."""
    if isinstance(valor, (list, tuple)):
        items = valor
    else:
        items = re.split(r"[,;/]+", str(valor or ""))
    out = []
    for item in items:
        nombre = re.sub(r"\s+", " ", str(item)).strip().upper()
        if nombre and nombre not in out:
            out.append(nombre)
    return out


@competidores_bp.route("/import", methods=["POST"])
@jwt_required()
@limitar(10, 60, nombre="competidores-import")
def importar_excel():
    """
    POST /api/competidores/import  (multipart/form-data)
    - file: hoja .xlsx (usa la plantilla o encabezados equivalentes).
    - campeonato_id (opcional): además de registrar, inscribe en el campeonato.
    - modalidades (opcional): por defecto para filas sin columna Modalidades,
      separadas por coma.
    Crea competidores nuevos y actualiza los existentes (por documento, o por
    nombre exacto si no hay documento).
    """
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    archivo = request.files.get("file")
    if not archivo or not archivo.filename:
        return jsonify({"error": "Adjunta el archivo Excel (.xlsx) en el campo 'file'."}), 400

    campeonato = None
    campeonato_id = request.form.get("campeonato_id")
    if campeonato_id:
        campeonato = Campeonato.query.get(int(campeonato_id))
        if not campeonato or not es_dueno_campeonato(admin, campeonato):
            return jsonify({"error": "Campeonato no encontrado"}), 404
    modalidades_default = _parse_modalidades(request.form.get("modalidades"))

    filas, error = _leer_excel(archivo)
    if error:
        return jsonify({"error": error}), 400
    if not filas:
        return jsonify({"error": "El archivo no tiene filas con datos."}), 400

    creados = actualizados = inscritos = 0
    errores = []

    for datos in filas:
        numero = datos.pop("_fila")
        nombre = str(datos.get("nombre_completo") or "").strip()
        if not nombre:
            errores.append({"fila": numero, "error": "Falta el nombre."})
            continue

        documento = str(datos.get("documento") or "").strip() or None
        # Workspace: el "match" de existentes se busca SOLO entre los
        # competidores del admin actual (superadmin busca en todos). Así un
        # admin no actualiza por accidente atletas de otro workspace.
        base = filtrar_competidores(admin, Competidor.query)
        comp = None
        if documento:
            comp = base.filter_by(documento=documento).first()
        if not comp:
            comp = base.filter(
                func.lower(Competidor.nombre_completo) == nombre.lower()
            ).first()
            # Si el registro homónimo tiene otro documento, es otra persona.
            if comp and documento and comp.documento and comp.documento != documento:
                comp = None

        es_nuevo = comp is None
        if es_nuevo:
            comp = Competidor(nombre_completo=nombre, activo=True, created_by=admin.id)
            db.session.add(comp)
        error = _aplicar_datos(comp, datos, parciales=True, actor=admin)
        if error:
            errores.append({"fila": numero, "error": error})
            if es_nuevo:
                db.session.expunge(comp)
            continue
        if es_nuevo:
            creados += 1
        else:
            actualizados += 1

        if campeonato:
            db.session.flush()
            ya = Inscripcion.query.filter_by(
                campeonato_id=campeonato.id, competidor_id=comp.id
            ).first()
            if not ya:
                modalidades = (
                    _parse_modalidades(datos.get("modalidades"))
                    or modalidades_default
                )
                db.session.add(Inscripcion(
                    campeonato_id=campeonato.id,
                    competidor_id=comp.id,
                    modalidades=modalidades or None,
                    peso=comp.peso,
                    grupo_cinturon=comp.grupo_cinturon,
                    estado="aceptada",
                    created_by=admin.id,
                ))
                inscritos += 1

    db.session.commit()
    return jsonify({
        "message": (
            f"Importación completada: {creados} nuevos, {actualizados} actualizados"
            + (f", {inscritos} inscritos" if campeonato else "")
            + (f", {len(errores)} con error" if errores else "")
        ),
        "creados": creados,
        "actualizados": actualizados,
        "inscritos": inscritos,
        "errores": errores,
    }), 200


# ══════════════════════════════════════════════════════════════════
#  Inscripciones (competidor ↔ campeonato)
# ══════════════════════════════════════════════════════════════════

@inscripciones_bp.route("/campeonato/<int:camp_id>", methods=["GET"])
@jwt_required()
def listar_inscripciones(camp_id):
    """
    GET /api/inscripciones/campeonato/:id?estado= — Inscritos del campeonato.
    `estado` opcional (aceptada|pendiente|rechazada) filtra la lista; sin él se
    devuelven todas (para que el admin vea las solicitudes por revisar).
    """
    from ..models.competidor import ESTADOS_INSCRIPCION

    camp = Campeonato.query.get_or_404(camp_id)
    user = usuario_actual()
    if user and user.rol == "admin" and not es_dueno_campeonato(user, camp):
        return jsonify({"error": "Campeonato no encontrado"}), 404
    query = Inscripcion.query.filter_by(campeonato_id=camp_id)
    estado = request.args.get("estado")
    if estado in ESTADOS_INSCRIPCION:
        query = query.filter_by(estado=estado)
    inscripciones = (
        query.join(Competidor)
        .order_by(Competidor.nombre_completo.asc())
        .all()
    )
    return jsonify([i.to_dict() for i in inscripciones]), 200


@inscripciones_bp.route("/campeonato/<int:camp_id>", methods=["POST"])
@jwt_required()
def inscribir(camp_id):
    """
    POST /api/inscripciones/campeonato/:id
    Body: {
      "competidor_id"?: int,        ← inscribir uno existente
      "competidor"?: { ... },       ← o crearlo en el sistema e inscribirlo
      "modalidades"?: ["COMBATE"], "peso"?: number
    }
    """
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    campeonato = Campeonato.query.get_or_404(camp_id)
    if not es_dueno_campeonato(admin, campeonato):
        return jsonify({"error": "Campeonato no encontrado"}), 404
    data = request.get_json() or {}

    comp = None
    if data.get("competidor_id"):
        comp = Competidor.query.get(int(data["competidor_id"]))
        if not comp or not es_dueno_competidor(admin, comp):
            return jsonify({"error": "Competidor no encontrado"}), 404
    elif isinstance(data.get("competidor"), dict):
        comp = Competidor(nombre_completo="", activo=True, created_by=admin.id)
        error = _aplicar_datos(comp, data["competidor"], actor=admin)
        if error:
            return jsonify({"error": error}), 400
        db.session.add(comp)
        db.session.flush()
    else:
        return jsonify({"error": "Indica competidor_id o los datos del competidor."}), 400

    ya = Inscripcion.query.filter_by(
        campeonato_id=campeonato.id, competidor_id=comp.id
    ).first()
    if ya:
        return jsonify({
            "error": f"{comp.nombre_completo} ya está inscrito en este campeonato."
        }), 409

    peso = None
    if data.get("peso") not in (None, ""):
        peso, error = _validar_peso(data.get("peso"))
        if error:
            return jsonify({"error": error}), 400
    inscripcion = Inscripcion(
        campeonato_id=campeonato.id,
        competidor_id=comp.id,
        modalidades=_parse_modalidades(data.get("modalidades")) or None,
        peso=peso if peso is not None else comp.peso,
        grupo_cinturon=comp.grupo_cinturon,
        # Las inscripciones del admin pasan directo, sin importar el estado
        # del campeonato.
        estado="aceptada",
        created_by=admin.id,
    )
    db.session.add(inscripcion)
    db.session.commit()
    return jsonify({
        "message": f"{comp.nombre_completo} inscrito en {campeonato.nombre}",
        "inscripcion": inscripcion.to_dict(),
    }), 201


@inscripciones_bp.route("/<int:ins_id>", methods=["PUT"])
@jwt_required()
def editar_inscripcion(ins_id):
    """PUT /api/inscripciones/:id — Modalidades / peso / cinturón de la inscripción."""
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    inscripcion = Inscripcion.query.get_or_404(ins_id)
    if not es_dueno_campeonato(admin, inscripcion.campeonato):
        return jsonify({"error": "Inscripción no encontrada"}), 404
    data = request.get_json() or {}

    if "modalidades" in data:
        inscripcion.modalidades = _parse_modalidades(data.get("modalidades")) or None
    if "peso" in data:
        if data.get("peso") in (None, ""):
            inscripcion.peso = None
        else:
            peso, error = _validar_peso(data.get("peso"))
            if error:
                return jsonify({"error": error}), 400
            inscripcion.peso = peso

    db.session.commit()
    return jsonify({
        "message": "Inscripción actualizada",
        "inscripcion": inscripcion.to_dict(),
    }), 200


@inscripciones_bp.route("/<int:ins_id>", methods=["DELETE"])
@jwt_required()
def eliminar_inscripcion(ins_id):
    """DELETE /api/inscripciones/:id — Quitar a un competidor del campeonato."""
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    inscripcion = Inscripcion.query.get_or_404(ins_id)
    if not es_dueno_campeonato(admin, inscripcion.campeonato):
        return jsonify({"error": "Inscripción no encontrada"}), 404
    nombre = inscripcion.competidor.nombre_completo if inscripcion.competidor else "?"
    db.session.delete(inscripcion)
    db.session.commit()
    return jsonify({"message": f"Inscripción de '{nombre}' eliminada"}), 200


@inscripciones_bp.route("/<int:ins_id>/estado", methods=["PATCH"])
@jwt_required()
def moderar_inscripcion(ins_id):
    """
    PATCH /api/inscripciones/:id/estado
    Body: { "estado": "aceptada" | "rechazada", "motivo"?: str }
    El admin dueño del campeonato acepta o rechaza una solicitud de un maestro.
    """
    admin = _require_admin()
    if not admin:
        return jsonify({"error": "Solo administradores"}), 403

    inscripcion = Inscripcion.query.get_or_404(ins_id)
    if not es_dueno_campeonato(admin, inscripcion.campeonato):
        return jsonify({"error": "Inscripción no encontrada"}), 404

    data = request.get_json() or {}
    nuevo = data.get("estado")
    if nuevo not in ("aceptada", "rechazada"):
        return jsonify({"error": "Estado inválido: usa 'aceptada' o 'rechazada'."}), 400

    inscripcion.estado = nuevo
    if nuevo == "rechazada":
        motivo = str(data.get("motivo") or "").strip()
        if len(motivo) > 300:
            return jsonify({"error": "El motivo no puede superar 300 caracteres."}), 400
        inscripcion.motivo_rechazo = motivo or None
    else:
        inscripcion.motivo_rechazo = None

    db.session.commit()
    nombre = inscripcion.competidor.nombre_completo if inscripcion.competidor else "?"
    verbo = "aceptada" if nuevo == "aceptada" else "rechazada"
    return jsonify({
        "message": f"Inscripción de '{nombre}' {verbo}",
        "inscripcion": inscripcion.to_dict(),
    }), 200


# ══════════════════════════════════════════════════════════════════
#  Flujo del MAESTRO (inscribe a sus alumnos → solicitud "pendiente")
# ══════════════════════════════════════════════════════════════════

@inscripciones_bp.route("/maestro/campeonatos", methods=["GET"])
@jwt_required()
def maestro_campeonatos():
    """
    GET /api/inscripciones/maestro/campeonatos
    Campeonatos activos del workspace del maestro con su estado. Solo puede
    inscribir en los que estén en 'preparacion' (puede_inscribir=True).
    """
    maestro = require_maestro()
    if not maestro:
        return jsonify({"error": "Solo maestros"}), 403

    query = filtrar_campeonatos(maestro, Campeonato.query.filter_by(activo=True))
    camps = query.order_by(Campeonato.created_at.desc()).all()
    return jsonify([{
        "id": c.id,
        "nombre": c.nombre,
        "descripcion": c.descripcion,
        "estado": c.estado or "preparacion",
        "fecha_inicio": c.fecha_inicio.isoformat() if c.fecha_inicio else None,
        "fecha_fin": c.fecha_fin.isoformat() if c.fecha_fin else None,
        "lugar": c.lugar,
        "ciudad": c.ciudad,
        "pais": c.pais,
        "puede_inscribir": (c.estado or "preparacion") == "preparacion",
    } for c in camps]), 200


@inscripciones_bp.route("/maestro/campeonato/<int:camp_id>", methods=["POST"])
@jwt_required()
def maestro_inscribir(camp_id):
    """
    POST /api/inscripciones/maestro/campeonato/:id
    Body: { "competidor": { ..., "club"? }, "modalidades"?: [...], "peso"?: number }
    El maestro envía una solicitud (queda 'pendiente'). El club del alumno tiene
    que ser uno de los suyos —si dirige varios, elige cuál; si no lo manda, se
    usa su club principal—; el alumno queda bajo el workspace de su admin.
    Solo si el campeonato está en 'preparacion' y es del workspace del maestro.
    """
    maestro = require_maestro()
    if not maestro:
        return jsonify({"error": "Solo maestros"}), 403

    campeonato = Campeonato.query.get_or_404(camp_id)
    if not es_dueno_campeonato(maestro, campeonato):
        return jsonify({"error": "Campeonato no encontrado"}), 404
    if (campeonato.estado or "preparacion") != "preparacion":
        return jsonify({
            "error": "Este campeonato ya no está en preparación: no acepta nuevas inscripciones."
        }), 409

    data = request.get_json() or {}
    datos = dict(data.get("competidor") or {})
    club, error = _club_del_maestro(maestro, datos.get("club"))
    if error:
        return jsonify({"error": error}), 400
    datos["club"] = club
    comp = Competidor(
        nombre_completo="", activo=True, created_by=workspace_owner_id(maestro)
    )
    error = _aplicar_datos(comp, datos, actor=maestro)
    if error:
        return jsonify({"error": error}), 400
    db.session.add(comp)
    db.session.flush()

    peso = None
    if data.get("peso") not in (None, ""):
        peso, error = _validar_peso(data.get("peso"))
        if error:
            return jsonify({"error": error}), 400
    inscripcion = Inscripcion(
        campeonato_id=campeonato.id,
        competidor_id=comp.id,
        modalidades=_parse_modalidades(data.get("modalidades")) or None,
        peso=peso if peso is not None else comp.peso,
        grupo_cinturon=comp.grupo_cinturon,
        estado="pendiente",
        created_by=maestro.id,
    )
    db.session.add(inscripcion)
    db.session.commit()
    return jsonify({
        "message": f"Solicitud enviada: {comp.nombre_completo}. El administrador la revisará.",
        "inscripcion": inscripcion.to_dict(),
    }), 201


@inscripciones_bp.route("/maestro/mias", methods=["GET"])
@jwt_required()
def maestro_mis_inscripciones():
    """
    GET /api/inscripciones/maestro/mias?campeonato_id=
    Solicitudes enviadas por el maestro actual, con su estado y motivo.
    """
    maestro = require_maestro()
    if not maestro:
        return jsonify({"error": "Solo maestros"}), 403

    query = Inscripcion.query.filter_by(created_by=maestro.id)
    camp_id = request.args.get("campeonato_id")
    if camp_id:
        try:
            query = query.filter_by(campeonato_id=int(camp_id))
        except (TypeError, ValueError):
            return jsonify({"error": "campeonato_id inválido"}), 400
    inscripciones = query.order_by(Inscripcion.created_at.desc()).all()
    return jsonify([i.to_dict() for i in inscripciones]), 200


@inscripciones_bp.route("/maestro/<int:ins_id>", methods=["PUT"])
@jwt_required()
def maestro_reenviar(ins_id):
    """
    PUT /api/inscripciones/maestro/:id
    Body: { "competidor": { ... }, "modalidades"?: [...] }
    Re-envía una inscripción RECHAZADA: actualiza los datos del competidor,
    restablece el estado a 'pendiente' y borra el motivo de rechazo.
    Solo funciona si:
      (a) la inscripción pertenece al maestro,
      (b) el estado actual es 'rechazada',
      (c) el campeonato sigue en 'preparacion'.
    """
    maestro = require_maestro()
    if not maestro:
        return jsonify({"error": "Solo maestros"}), 403

    inscripcion = Inscripcion.query.get_or_404(ins_id)
    if inscripcion.created_by != maestro.id:
        return jsonify({"error": "Inscripción no encontrada"}), 404

    if inscripcion.estado != "rechazada":
        return jsonify({
            "error": "Solo puedes corregir inscripciones rechazadas."
        }), 409

    campeonato = Campeonato.query.get(inscripcion.campeonato_id)
    if not campeonato or (campeonato.estado or "preparacion") != "preparacion":
        return jsonify({
            "error": "Este campeonato ya no está en preparación: no acepta correcciones."
        }), 409

    data = request.get_json() or {}
    datos = dict(data.get("competidor") or {})

    # El club tiene que ser uno de los suyos (ver `_club_del_maestro`). Al
    # corregir un rechazo también puede cambiarlo: a veces el rechazo es
    # justamente porque el alumno se envió con el dojang equivocado.
    club, error = _club_del_maestro(maestro, datos.get("club"))
    if error:
        return jsonify({"error": error}), 400
    datos["club"] = club

    comp = inscripcion.competidor
    if not comp:
        return jsonify({"error": "Competidor no encontrado"}), 404

    error = _aplicar_datos(comp, datos, actor=maestro)
    if error:
        return jsonify({"error": error}), 400

    # Actualizar modalidades si se enviaron
    if "modalidades" in data:
        inscripcion.modalidades = _parse_modalidades(data.get("modalidades")) or None

    # Actualizar snapshot de peso y cinturón
    inscripcion.peso = comp.peso
    inscripcion.grupo_cinturon = comp.grupo_cinturon

    # Restablecer estado
    inscripcion.estado = "pendiente"
    inscripcion.motivo_rechazo = None

    db.session.commit()
    return jsonify({
        "message": f"Solicitud re-enviada: {comp.nombre_completo}. El administrador la revisará.",
        "inscripcion": inscripcion.to_dict(),
    }), 200


# ══════════════════════════════════════════════════════════════════
#  Ficha PÚBLICA del campeonato (sin login)
# ══════════════════════════════════════════════════════════════════

@inscripciones_bp.route("/publico/campeonato/<int:camp_id>", methods=["GET"])
@limitar(60, 60, nombre="inscripciones-publico")
def publico_campeonato(camp_id):
    """
    GET /api/inscripciones/publico/campeonato/:id — Sin login.
    Ficha pública: datos del campeonato + inscritos ACEPTADOS (nombre, club,
    modalidades) + jueces asignados + tatamis. Solo campeonatos activos.
    """
    from ..models.asignacion import AsignacionJuez
    from ..models.tatami import Tatami

    camp = Campeonato.query.filter_by(id=camp_id, activo=True).first()
    if not camp:
        return jsonify({"error": "Campeonato no encontrado"}), 404

    inscripciones = (
        Inscripcion.query.filter_by(campeonato_id=camp.id, estado="aceptada")
        .join(Competidor)
        .order_by(Competidor.nombre_completo.asc())
        .all()
    )
    competidores = [{
        "nombre": i.competidor.nombre_completo,
        "club": i.competidor.club or "",
        "modalidades": i.modalidades or [],
    } for i in inscripciones if i.competidor]

    tatamis = (
        Tatami.query.filter_by(campeonato_id=camp.id)
        .order_by(Tatami.numero.asc())
        .all()
    )
    tnum = {t.id: t.numero for t in tatamis}
    jueces = []
    if tnum:
        asigs = AsignacionJuez.query.filter(
            AsignacionJuez.tatami_id.in_(list(tnum.keys()))
        ).all()
        jueces = [{
            "nombre": a.nombre_display,
            "rol_tatami": a.rol_tatami,
            "tatami_numero": tnum.get(a.tatami_id),
        } for a in asigs]

    return jsonify({
        "campeonato": {
            "id": camp.id,
            "nombre": camp.nombre,
            "descripcion": camp.descripcion,
            "estado": camp.estado or "preparacion",
            "fecha_inicio": camp.fecha_inicio.isoformat() if camp.fecha_inicio else None,
            "fecha_fin": camp.fecha_fin.isoformat() if camp.fecha_fin else None,
            "lugar": camp.lugar,
            "ciudad": camp.ciudad,
            "pais": camp.pais,
        },
        "competidores": competidores,
        "jueces": jueces,
        "clubes": sorted({c["club"] for c in competidores if c["club"]}, key=str.lower),
        # La ficha pública muestra los tatamis del evento (y enlaza a su
        # pantalla cuando el campeonato ya está en curso).
        "tatamis": [
            {"id": t.id, "numero": t.numero, "activo": bool(t.activo)}
            for t in tatamis
        ],
    }), 200
